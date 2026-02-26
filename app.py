import streamlit as st
import sqlite3, httpx, json, re, random
from bs4 import BeautifulSoup
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

st.set_page_config(page_title="Amazon Analyser", page_icon="\U0001f6d2", layout="wide")

DARK_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');

/* ── RESET & BASE ── */
html, body, [class*="css"] { font-family: 'Inter', sans-serif !important; }
#MainMenu, footer, header { visibility: hidden; }
.stDeployButton { display: none !important; }

/* ── APP BACKGROUND ── */
.stApp { background: #0B0B15 !important; }
.main .block-container { max-width: 1080px; padding: 2.5rem 2.5rem 6rem; }

/* ── SIDEBAR ── */
[data-testid="stSidebar"] {
    background: #080810 !important;
    border-right: 1px solid rgba(124,58,237,.2) !important;
}
[data-testid="stSidebar"] * { color: #9090C0 !important; }
[data-testid="stSidebar"] .stRadio label {
    font-size: .875rem !important; font-weight: 500 !important;
    color: #A0A0D0 !important; padding: 9px 14px !important;
    border-radius: 9px !important; display: block !important;
    transition: all .15s !important; margin: 2px 0 !important;
}
[data-testid="stSidebar"] .stRadio label:hover {
    background: rgba(124,58,237,.15) !important; color: #C8B0FF !important;
}
[data-testid="stSidebar"] hr { border-color: rgba(124,58,237,.2) !important; }

/* ── METRIC CARDS ── */
[data-testid="metric-container"] {
    background: #12122A !important;
    border: 1px solid rgba(124,58,237,.3) !important;
    border-radius: 18px !important; padding: 22px 26px !important;
    box-shadow: 0 0 30px rgba(124,58,237,.07) !important;
}
[data-testid="metric-container"] label {
    font-size: .7rem !important; font-weight: 700 !important;
    text-transform: uppercase !important; letter-spacing: .08em !important;
    color: #7B7BAA !important;
}
[data-testid="metric-container"] [data-testid="stMetricValue"] {
    font-size: 2.3rem !important; font-weight: 800 !important;
    color: #F0F0FF !important; letter-spacing: -.04em !important;
}

/* ── BUTTONS ── */
.stButton > button {
    font-family: 'Inter', sans-serif !important; font-weight: 600 !important;
    font-size: .855rem !important; border-radius: 10px !important;
    padding: .5rem 1.4rem !important;
    border: 1px solid rgba(124,58,237,.3) !important;
    background: rgba(124,58,237,.1) !important;
    color: #C4B5FD !important; transition: all .18s !important;
    box-shadow: none !important;
}
.stButton > button:hover {
    border-color: rgba(124,58,237,.7) !important;
    background: rgba(124,58,237,.22) !important;
    color: #EDE9FE !important; transform: translateY(-1px) !important;
    box-shadow: 0 4px 16px rgba(124,58,237,.25) !important;
}
.stButton > button[kind="primary"] {
    background: linear-gradient(135deg, #6D28D9, #8B5CF6) !important;
    color: #FFFFFF !important; border: none !important;
    box-shadow: 0 4px 20px rgba(109,40,217,.4) !important;
}
.stButton > button[kind="primary"]:hover {
    background: linear-gradient(135deg, #5B21B6, #7C3AED) !important;
    box-shadow: 0 6px 28px rgba(109,40,217,.55) !important;
    color: #FFFFFF !important; border: none !important;
}

/* ── TEXT INPUTS ── */
[data-testid="stTextInput"] input,
[data-testid="stNumberInput"] input {
    background: #0F0F20 !important;
    border: 1.5px solid rgba(124,58,237,.25) !important;
    border-radius: 10px !important;
    color: #E8E8FF !important;
    font-size: .875rem !important;
    caret-color: #8B5CF6 !important;
}
[data-testid="stTextInput"] input:focus,
[data-testid="stNumberInput"] input:focus {
    border-color: #7C3AED !important;
    box-shadow: 0 0 0 3px rgba(124,58,237,.2) !important;
    outline: none !important;
}
[data-testid="stTextInput"] input::placeholder,
[data-testid="stNumberInput"] input::placeholder {
    color: #5A5A80 !important;
}

/* ── TEXT AREA ── */
[data-testid="stTextArea"] textarea {
    background: #0F0F20 !important;
    border: 1.5px solid rgba(124,58,237,.25) !important;
    border-radius: 12px !important;
    color: #E8E8FF !important;
    font-size: .875rem !important; line-height: 1.65 !important;
    caret-color: #8B5CF6 !important;
}
[data-testid="stTextArea"] textarea:focus {
    border-color: #7C3AED !important;
    box-shadow: 0 0 0 3px rgba(124,58,237,.2) !important;
}
[data-testid="stTextArea"] textarea::placeholder { color: #5A5A80 !important; }
[data-testid="stTextArea"] label { color: #A0A0D0 !important; font-size: .8rem !important; }

/* ── SELECTBOX — THE MAIN FIX ── */
[data-testid="stSelectbox"] > div > div {
    background: #0F0F20 !important;
    border: 1.5px solid rgba(124,58,237,.25) !important;
    border-radius: 10px !important;
    color: #E8E8FF !important;
}
[data-testid="stSelectbox"] > div > div > div { color: #E8E8FF !important; }
[data-testid="stSelectbox"] svg { fill: #8B5CF6 !important; }
/* Dropdown popup */
[data-baseweb="popover"], [data-baseweb="menu"],
div[role="listbox"], ul[role="listbox"] {
    background: #1A1A2E !important;
    border: 1px solid rgba(124,58,237,.35) !important;
    border-radius: 10px !important;
}
li[role="option"] {
    background: #1A1A2E !important; color: #E8E8FF !important;
    font-size: .875rem !important;
}
li[role="option"]:hover, li[aria-selected="true"] {
    background: rgba(124,58,237,.2) !important; color: #EDE9FE !important;
}

/* ── EXPANDER ── */
[data-testid="stExpander"] {
    background: #10101E !important;
    border: 1px solid rgba(124,58,237,.18) !important;
    border-radius: 16px !important; margin-bottom: 10px !important;
    overflow: hidden !important;
}
[data-testid="stExpander"] summary {
    font-weight: 600 !important; font-size: .88rem !important;
    color: #C8C8F0 !important; padding: 16px 20px !important;
    background: #10101E !important; letter-spacing: -.01em !important;
    border-radius: 16px !important;
}
[data-testid="stExpander"] summary:hover { background: #16162A !important; }
[data-testid="stExpander"] > div:last-child {
    padding: 4px 20px 20px !important; background: #10101E !important;
}
[data-testid="stExpander"] summary svg { fill: #8B5CF6 !important; }

/* ── NUMBER INPUT STEPPERS ── */
[data-testid="stNumberInput"] button {
    background: #1A1A2E !important; color: #8B5CF6 !important;
    border: 1px solid rgba(124,58,237,.2) !important;
    border-radius: 8px !important;
}

/* ── CHECKBOX ── */
[data-testid="stCheckbox"] label { color: #A0A0D0 !important; font-size: .875rem !important; }
[data-testid="stCheckbox"] span { border-color: rgba(124,58,237,.4) !important; }

/* ── SLIDER ── */
[data-testid="stSlider"] { color: #A0A0D0 !important; }
[data-testid="stSlider"] > div > div > div { background: #7C3AED !important; }
[data-testid="stSlider"] label { color: #A0A0D0 !important; }

/* ── ALERTS ── */
[data-testid="stAlert"] {
    background: rgba(239,68,68,.1) !important;
    border: 1px solid rgba(239,68,68,.3) !important;
    border-radius: 12px !important; color: #FCA5A5 !important;
    font-size: .855rem !important;
}

/* ── PROGRESS BAR ── */
[data-testid="stProgressBar"] > div { background: rgba(255,255,255,.07) !important; border-radius: 999px !important; }
[data-testid="stProgressBar"] > div > div { background: linear-gradient(90deg,#7C3AED,#8B5CF6) !important; border-radius: 999px !important; }

/* ── DOWNLOAD BUTTON ── */
[data-testid="stDownloadButton"] > button {
    background: rgba(124,58,237,.1) !important;
    border: 1px solid rgba(124,58,237,.3) !important;
    color: #C4B5FD !important; border-radius: 10px !important;
    font-weight: 600 !important;
}
[data-testid="stDownloadButton"] > button:hover {
    background: rgba(124,58,237,.22) !important;
    border-color: rgba(124,58,237,.6) !important; color: #EDE9FE !important;
}

/* ── CAPTION / SMALL TEXT ── */
[data-testid="stCaptionContainer"] { color: #7B7BAA !important; font-size: .78rem !important; }

/* ── DIVIDER ── */
hr { border: none !important; border-top: 1px solid rgba(255,255,255,.07) !important; margin: 1.5rem 0 !important; }

/* ── INLINE CODE ── */
code {
    background: rgba(124,58,237,.2) !important; color: #C4B5FD !important;
    border-radius: 6px !important; padding: 2px 8px !important;
    font-size: .82rem !important; font-family: 'SF Mono','Fira Code',monospace !important;
}

/* ── BASE TEXT ── */
p, li { color: #A0A0D0 !important; line-height: 1.65 !important; }
h1, h2, h3 { color: #F0F0FF !important; }

/* ── RADIO GROUP CONTAINER ── */
[data-testid="stRadio"] > label { color: #7B7BAA !important; font-size: .72rem !important; font-weight: 700 !important; text-transform: uppercase !important; letter-spacing: .07em !important; }

/* ── COLUMNS GAP FIX ── */
[data-testid="column"] { padding-left: 8px !important; padding-right: 8px !important; }

/* ── SCROLLBAR ── */
::-webkit-scrollbar { width: 6px; height: 6px; }
::-webkit-scrollbar-track { background: #0B0B15; }
::-webkit-scrollbar-thumb { background: rgba(124,58,237,.3); border-radius: 3px; }
::-webkit-scrollbar-thumb:hover { background: rgba(124,58,237,.5); }
</style>
"""
st.markdown(DARK_CSS, unsafe_allow_html=True)


# ── DATABASE ───────────────────────────────────────────────────────────────────
DB = "amazon_analyser.db"

def db():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    with db() as c:
        c.executescript("""
        CREATE TABLE IF NOT EXISTS product_lines (
            id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT UNIQUE NOT NULL,
            keywords TEXT DEFAULT '[]', created_at TEXT DEFAULT (datetime('now')));
        CREATE TABLE IF NOT EXISTS scoring_rules (
            id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT UNIQUE NOT NULL,
            title_weight REAL DEFAULT 25, bullets_weight REAL DEFAULT 25,
            aplus_weight REAL DEFAULT 25, keywords_weight REAL DEFAULT 25,
            title_min_length INTEGER DEFAULT 80, title_max_length INTEGER DEFAULT 200,
            title_keyword_in_first INTEGER DEFAULT 1, bullets_min_count INTEGER DEFAULT 5,
            bullets_min_length INTEGER DEFAULT 100, bullets_max_length INTEGER DEFAULT 255,
            keywords_min_coverage REAL DEFAULT 70, keywords_in_title INTEGER DEFAULT 1,
            is_default INTEGER DEFAULT 0, created_at TEXT DEFAULT (datetime('now')));
        CREATE TABLE IF NOT EXISTS results (
            id INTEGER PRIMARY KEY AUTOINCREMENT, asin TEXT NOT NULL,
            product_line_id INTEGER, scoring_rule_id INTEGER,
            product_name TEXT, product_image TEXT, title TEXT,
            bullets TEXT DEFAULT '[]', has_aplus INTEGER DEFAULT 0, description TEXT,
            scrape_error TEXT, total_score REAL,
            title_score REAL, bullets_score REAL, aplus_score REAL, keywords_score REAL,
            title_issues TEXT DEFAULT '[]', bullets_issues TEXT DEFAULT '[]',
            aplus_issues TEXT DEFAULT '[]', keywords_issues TEXT DEFAULT '[]',
            found_keywords TEXT DEFAULT '[]', missing_keywords TEXT DEFAULT '[]',
            suggested_keywords TEXT DEFAULT '[]', created_at TEXT DEFAULT (datetime('now')));
        """)
        if not c.execute("SELECT id FROM scoring_rules WHERE is_default=1").fetchone():
            c.execute("INSERT INTO scoring_rules (name,is_default) VALUES ('Default Rule',1)")

init_db()

# ── A+ DETECTION — 7 strategies ───────────────────────────────────────────────
def detect_aplus(soup, html_text):
    for aid in ["aplus","aplus3p_feature_div","aplusBrandStory_feature_div",
                "ap-desktop-product-description-feature-div","aplus_feature_div",
                "aplusWidget","aplus-tech-specs","dpx_aplus_description_feature_div"]:
        el = soup.find(attrs={"id": aid})
        if el and len(el.get_text(strip=True)) > 30:
            return True
    for cls in ["aplus-v2","aplus-module","a-aplus","aplusBrandStory",
                "aplus-brand-story-hero","aplus-brand-story-card","ap-comparison-table"]:
        found = soup.find(attrs={"class": lambda c, _c=cls: c and _c in (c if isinstance(c,str) else " ".join(c))})
        if found and len(found.get_text(strip=True)) > 30:
            return True
    for el in soup.find_all(attrs={"data-feature-name": True}):
        fn = el.get("data-feature-name","").lower()
        if any(s in fn for s in ["aplus","aplusbrandstory","dpx_aplus"]):
            if len(el.get_text(strip=True)) > 30: return True
    for el in soup.find_all(attrs={"data-cel-widget": True}):
        if "aplus" in el.get("data-cel-widget","").lower(): return True
    hl = html_text.lower()
    for sig in ['"aplus-v2"',"aplus-module","aplusbrandstory","aplus3p_feature_div",
                "a-aplus","enhanced-brand-content","aplus_feature","aplus-tech-specs"]:
        if sig in hl: return True
    for img in soup.find_all("img", src=True):
        src = img.get("src","").lower()
        if "aplus" in src or "a-plus" in src: return True
    for script in soup.find_all("script"):
        if script.string and "aplus" in script.string.lower(): return True
    return False

# ── SCRAPER ────────────────────────────────────────────────────────────────────
UA = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/122.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/122.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:124.0) Gecko/20100101 Firefox/124.0",
]

def scrape(asin):
    url = "https://www.amazon.co.uk/dp/" + asin
    hdrs = {"User-Agent": random.choice(UA), "Accept": "text/html,application/xhtml+xml,*/*;q=0.8",
            "Accept-Language": "en-GB,en;q=0.9", "Accept-Encoding": "gzip, deflate, br",
            "Connection": "keep-alive", "Upgrade-Insecure-Requests": "1"}
    with httpx.Client(follow_redirects=True, timeout=30.0) as client:
        r = client.get(url, headers=hdrs)
    if r.status_code == 503:
        raise Exception("Amazon returned 503 — rate limited. Wait a few minutes.")
    if r.status_code != 200:
        raise Exception("HTTP " + str(r.status_code) + " for ASIN " + asin)
    if "captcha" in r.text.lower() and "robot" in r.text.lower():
        raise Exception("Amazon returned a CAPTCHA. Try again in a few minutes.")
    soup = BeautifulSoup(r.text, "html.parser")
    title_el = soup.find("span", id="productTitle")
    title = title_el.get_text(strip=True) if title_el else None
    bullets = []
    feat = soup.find("div", id="feature-bullets")
    if feat:
        bullets = [i.get_text(strip=True) for i in feat.find_all("span", class_="a-list-item")
                   if len(i.get_text(strip=True)) > 10]
    has_aplus = detect_aplus(soup, r.text)
    desc_el = soup.find("div", id="productDescription")
    description = desc_el.get_text(strip=True) if desc_el else ""
    img_el = soup.find("img", id="landingImage") or soup.find("img", id="imgBlkFront")
    image_url = img_el.get("src") if img_el else None
    brand_el = soup.find("a", id="bylineInfo")
    brand = brand_el.get_text(strip=True) if brand_el else None
    return {"title": title, "bullets": bullets, "has_aplus": has_aplus,
            "description": description, "image_url": image_url, "brand": brand}


# ── ANALYSER ───────────────────────────────────────────────────────────────────
def kw_split(content, kws):
    cl = content.lower()
    return [k for k in kws if k.lower() in cl], [k for k in kws if k.lower() not in cl]

def analyse(scraped, keywords, rule):
    title = scraped.get("title") or ""
    bullets = scraped.get("bullets") or []
    has_aplus = scraped.get("has_aplus", False)
    desc = scraped.get("description") or ""
    mw = rule["title_weight"]; ti, ts = [], 0
    if not title:
        ti = ["Title could not be retrieved from Amazon."]
    else:
        ln = len(title)
        if ln < rule["title_min_length"]:
            ti.append("Title too short (" + str(ln) + " chars). Target: " + str(rule["title_min_length"]) + "\u2013" + str(rule["title_max_length"]) + " chars.")
            ts += mw * 0.4 * (ln / rule["title_min_length"])
        elif ln > rule["title_max_length"]:
            ti.append("Title too long (" + str(ln) + " chars) \u2014 Amazon may truncate it.")
            ts += mw * 0.30
        else:
            ts += mw * 0.40
        found_t, _ = kw_split(title, keywords)
        kr = len(found_t) / len(keywords) if keywords else 1.0
        ts += mw * 0.40 * kr
        if kr < 0.3: ti.append("Only " + str(len(found_t)) + "/" + str(len(keywords)) + " keywords in title.")
        elif kr < 0.6: ti.append(str(len(found_t)) + "/" + str(len(keywords)) + " keywords in title \u2014 room to improve.")
        if rule.get("title_keyword_in_first") and keywords:
            if not any(k.lower() in title[:80].lower() for k in keywords[:3]):
                ti.append("Primary keyword not in first 80 characters of the title.")
            else:
                ts += mw * 0.20
        if sum(1 for w in title.split() if w.isupper() and len(w) > 2) > 3:
            ti.append("Too many ALL-CAPS words \u2014 use Title Case.")
    mw2 = rule["bullets_weight"]; bi, bs = [], 0
    if not bullets:
        bi = ["No bullet points found on the product page."]
    else:
        cnt = len(bullets)
        if cnt < rule["bullets_min_count"]:
            bi.append("Only " + str(cnt) + " bullet(s). Recommended: " + str(rule["bullets_min_count"]) + ".")
            bs += mw2 * 0.30 * (cnt / rule["bullets_min_count"])
        else:
            bs += mw2 * 0.30
        short = [i+1 for i,b in enumerate(bullets) if len(b) < rule["bullets_min_length"]]
        long_ = [i+1 for i,b in enumerate(bullets) if len(b) > rule["bullets_max_length"]]
        if short: bi.append("Bullet(s) " + str(short) + " too short (< " + str(rule["bullets_min_length"]) + " chars).")
        if long_: bi.append("Bullet(s) " + str(long_) + " exceed " + str(rule["bullets_max_length"]) + " chars.")
        ok = (cnt - len(short) - len(long_)) / cnt
        bs += mw2 * 0.30 * ok
        found_b, _ = kw_split(" ".join(bullets), keywords)
        kr2 = len(found_b) / len(keywords) if keywords else 1.0
        bs += mw2 * 0.40 * kr2
        if kr2 < 0.5: bi.append("Low keyword coverage in bullets (" + str(len(found_b)) + "/" + str(len(keywords)) + ").")
    a_s = rule["aplus_weight"] if has_aplus else 0
    ai = [] if has_aplus else ["No A+ Content detected. Adding A+ can boost conversion by 5\u201310%."]
    mw4 = rule["keywords_weight"]; ki, ks = [], 0
    all_content = " ".join(filter(None, [title, " ".join(bullets), desc]))
    if not keywords:
        ks, found_k, miss_k = mw4, [], []
        ki = ["No keywords assigned to this product line."]
    else:
        found_k, miss_k = kw_split(all_content, keywords)
        cov = len(found_k) / len(keywords) * 100
        ks = mw4 * (cov / 100)
        thresh = rule.get("keywords_min_coverage", 70)
        if cov < thresh:
            ki.append("Coverage " + str(round(cov)) + "% (target \u2265 " + str(round(thresh)) + "%). Missing: " + ", ".join(miss_k[:8]) + ("..." if len(miss_k) > 8 else "") + ".")
        if rule.get("keywords_in_title") and keywords:
            ft2, _ = kw_split(title, keywords[:3])
            if not ft2: ki.append("None of the top 3 keywords in title.")
    tokens = re.findall(r"\b[a-zA-Z]{4,}\b", all_content.lower())
    stop = {"with","this","that","from","have","will","your","their","also","each","which","they","more","than","when","into","only","over","such","used","using","pack","item","product","brand","quality","great","make","made","features","feature","design","provides","include","perfect","ideal","easy","best","good","high","well","help","helps","allows","keep"}
    existing_l = {k.lower() for k in keywords}
    counts = {}
    for t in tokens:
        if t not in stop and t not in existing_l: counts[t] = counts.get(t, 0) + 1
    suggestions = sorted(counts, key=lambda x: -counts[x])[:15]
    total = round(min(ts,mw) + min(bs,mw2) + a_s + min(ks,mw4), 1)
    return {"total_score": total,
            "title_score": round(min(ts,mw),1), "bullets_score": round(min(bs,mw2),1),
            "aplus_score": round(a_s,1), "keywords_score": round(min(ks,mw4),1),
            "title_issues": ti, "bullets_issues": bi, "aplus_issues": ai, "keywords_issues": ki,
            "found_keywords": found_k, "missing_keywords": miss_k, "suggested_keywords": suggestions}

def keyword_gaps(title, bullets_list, pl_keywords):
    tl = (title or "").lower()
    bt = " ".join(bullets_list or []).lower()
    return [k for k in pl_keywords if k.lower() not in tl], [k for k in pl_keywords if k.lower() not in bt]

# ── SMART BULLET REWRITE ENGINE ────────────────────────────────────────────────
def classify_bullet(bullet):
    b = bullet.strip().lower()
    if re.search(r"\b\d+[\.,]?\d*\s*(mm|cm|m|km|kg|g|lb|oz|w|v|mah|l|ml|inch|ft|gb|tb|mhz|ghz|rpm|%)", b):
        return "specification"
    for v in ["features","includes","comes with","equipped with","built-in","contains","provides","offers"]:
        if b.startswith(v) or v in b[:50]: return "feature_list"
    for s in ["enjoy","experience","get ","achieve","maximize","boost","improve","enhance","ensure","guarantee","discover"]:
        if b.startswith(s): return "benefit"
    if any(q in b for q in ["premium","high-quality","durable","robust","professional","heavy-duty","lightweight","compact","ergonomic"]):
        return "quality"
    for s in ["easy to","simple to","quick to","simply ","just ","effortlessly","one-touch","one-click"]:
        if b.startswith(s): return "action"
    for c in ["compatible with","works with","suitable for","designed for","ideal for","perfect for","great for"]:
        if c in b: return "compatibility"
    for p in ["prevents","stops","protects","eliminates","reduces","no more","say goodbye"]:
        if b.startswith(p) or p in b[:40]: return "problem_solution"
    return "general"

def smart_rewrite(bullet, keyword):
    b = bullet.strip().rstrip(".")
    kw = keyword.strip(); kc = kw.capitalize(); bl = b.lower(); words = b.split()
    btype = classify_bullet(b)
    if kw.lower() in bl: return b + ".", "already-present", 100
    if btype == "feature_list":
        for verb in ["features","includes","comes with","equipped with"]:
            if verb in bl:
                idx = bl.index(verb) + len(verb); rest = b[idx:].strip()
                if "," in rest:
                    parts = rest.rsplit(",",1)
                    return b[:idx] + " " + parts[0] + ", " + kw + "," + parts[1] + ".", "feature-list-insert", 88
                if " and " in rest:
                    ai = rest.rindex(" and ")
                    return b[:idx] + " " + rest[:ai] + ", " + kw + rest[ai:] + ".", "feature-list-and", 86
                return b[:idx] + " " + rest + " and " + kw + ".", "feature-list-extend", 82
        return b + " with " + kw + ".", "feature-list-suffix", 72
    if btype == "benefit":
        if len(words) >= 3: return words[0] + " " + kw + " " + " ".join(words[1:]) + ".", "benefit-subject", 84
        return kc + " \u2014 " + b + ".", "benefit-prefix", 72
    if btype == "specification": return kc + ": " + b + ".", "spec-qualifier", 85
    if btype == "quality":
        for qa in ["premium","durable","professional","high-quality","robust","lightweight","compact"]:
            if qa in bl:
                qi = bl.index(qa) + len(qa)
                return b[:qi] + " " + kw + b[qi:] + ".", "quality-adjective", 83
        return kc + " \u2014 " + b + ".", "quality-prefix", 70
    if btype == "action": return b + " for " + kw + " use.", "action-usecase", 79
    if btype == "compatibility":
        for phrase in ["ideal for","perfect for","suitable for","designed for"]:
            if phrase in bl:
                pi = bl.index(phrase) + len(phrase); rest = b[pi:].strip()
                if rest: return b[:pi] + " " + kw + " and " + rest + ".", "compat-extend", 87
                return b[:pi] + " " + kw + ".", "compat-add", 82
        return b + ", " + kw + " compatible.", "compat-suffix", 74
    if btype == "problem_solution": return b + " \u2014 ideal for " + kw + ".", "problem-ideal", 74
    if " and " in bl:
        ai = bl.rindex(" and ")
        if ai > 5: return b[:ai] + ", " + kw + b[ai:] + ".", "general-and-insert", 74
    if b.count(",") >= 2:
        parts = b.rsplit(",",1)
        return parts[0] + ", " + kw + "," + parts[1] + ".", "general-comma-insert", 77
    if len(b) < 70: return kc + " \u2014 " + b + ".", "general-category-prefix", 70
    for sep in ["; "," \u2014 "," - ",": "]:
        if sep in b:
            si = b.index(sep); rest = b[si+len(sep):]
            return b[:si] + sep + kc + " " + rest + ".", "general-clause-split", 71
    return b + " \u2014 includes " + kw + ".", "general-append", 65

def bullet_rewrite_suggestions(bullets_list, missing_kws):
    suggestions = []
    for kw in missing_kws:
        if not bullets_list: break
        kw_words = set(re.findall(r"\b\w{3,}\b", kw.lower()))
        best_idx, best_score = 0, -1
        for i, bul in enumerate(bullets_list):
            bul_words = set(re.findall(r"\b\w{3,}\b", bul.lower()))
            score = (len(kw_words & bul_words) * 3 + max(0,(300-len(bul))/300)
                     + (0.5 if "," in bul or " and " in bul.lower() else 0))
            if score > best_score: best_score, best_idx = score, i
        orig = bullets_list[best_idx]
        sugg, strat, conf = smart_rewrite(orig, kw)
        suggestions.append({"keyword": kw, "bullet_idx": best_idx,
                             "bullet_type": classify_bullet(orig),
                             "original": orig, "suggested": sugg,
                             "strategy": strat, "confidence": conf})
    return suggestions

# ── EXCEL EXPORT ───────────────────────────────────────────────────────────────
def build_excel(rows):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Analysis Results"
    ws.freeze_panes = "A2"
    hf    = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    hfill = PatternFill("solid", fgColor="12122A")
    alt   = PatternFill("solid", fgColor="0B0B15")
    thin  = Border(bottom=Side(style="thin", color="1E1E3A"))
    hdrs  = ["ASIN","Product","Date","Score","Title","Bullets","A+","Keywords",
             "Has A+","Title Missing KWs","Bullets Missing KWs","All Missing KWs","Issues"]
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hf; cell.fill = hfill; cell.border = thin
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for ri, r in enumerate(rows, 2):
        blist  = json.loads(r["bullets"] or "[]")
        pl_kws = json.loads(r["pl_keywords"] or "[]") if "pl_keywords" in r.keys() else []
        tm, bm = keyword_gaps(r["title"] or "", blist, pl_kws) if pl_kws else ([], [])
        vals = [r["asin"], r["product_name"] or "\u2014", str(r["created_at"] or "")[:10],
                r["total_score"], r["title_score"], r["bullets_score"], r["aplus_score"], r["keywords_score"],
                "Yes" if r["has_aplus"] else "No",
                "; ".join(tm), "; ".join(bm),
                "; ".join(json.loads(r["missing_keywords"] or "[]")),
                "; ".join(json.loads(r["title_issues"] or "[]") + json.loads(r["bullets_issues"] or "[]") +
                           json.loads(r["aplus_issues"] or "[]") + json.loads(r["keywords_issues"] or "[]"))]
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=c, value=val); cell.border = thin
            cell.alignment = Alignment(vertical="top", wrap_text=(c >= 10))
            if ri % 2 == 0: cell.fill = alt
            if c == 4 and val is not None:
                v = float(val)
                col = ("16A34A" if v>=85 else "CA8A04" if v>=65 else "EA580C" if v>=40 else "DC2626")
                cell.font = Font(bold=True, color=col, name="Calibri", size=10)
            else:
                cell.font = Font(name="Calibri", size=10)
    for i, w in enumerate([14,28,12,10,10,10,8,10,8,30,30,30,50], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 20
    out = BytesIO(); wb.save(out); out.seek(0)
    return out.getvalue()


# ── UI HELPERS ─────────────────────────────────────────────────────────────────
def _score_meta(s):
    if s is None: s = 0
    if s >= 85: return "#22C55E","rgba(34,197,94,.15)","#86EFAC","Excellent"
    if s >= 65: return "#EAB308","rgba(234,179,8,.15)","#FDE047","Good"
    if s >= 40: return "#F97316","rgba(249,115,22,.15)","#FED7AA","Needs Work"
    return "#EF4444","rgba(239,68,68,.15)","#FCA5A5","Poor"

def _bar_c(pct):
    if pct >= 85: return "#22C55E"
    if pct >= 65: return "#EAB308"
    if pct >= 40: return "#F97316"
    return "#EF4444"

def show(html): st.markdown(html, unsafe_allow_html=True)

def score_gauge(s, size=112):
    if s is None: s = 0
    c, bg, tc, lbl = _score_meta(s)
    pct = max(0, min(100, s)); rem = 100 - pct
    return (
        '<div style="display:flex;flex-direction:column;align-items:center;gap:10px">'
        '<div style="position:relative;width:' + str(size) + 'px;height:' + str(size) + 'px">'
        '<svg viewBox="0 0 36 36" style="width:100%;height:100%;transform:rotate(-90deg)">'
        '<circle cx="18" cy="18" r="15.9155" fill="none" stroke="rgba(255,255,255,.06)" stroke-width="3"/>'
        '<circle cx="18" cy="18" r="15.9155" fill="none" stroke="' + c + '" stroke-width="3"'
        ' stroke-dasharray="' + str(round(pct,1)) + ' ' + str(round(rem,1)) + '"'
        ' stroke-linecap="round" style="filter:drop-shadow(0 0 8px ' + c + ')"/>'
        '</svg>'
        '<div style="position:absolute;inset:0;display:flex;flex-direction:column;align-items:center;justify-content:center">'
        '<span style="font-size:1.6rem;font-weight:800;color:#F0F0FF;line-height:1;letter-spacing:-.04em">' + str(int(s)) + '</span>'
        '<span style="font-size:.6rem;color:#7B7BAA;font-weight:600;letter-spacing:.05em;margin-top:2px">/ 100</span>'
        '</div></div>'
        '<span style="font-size:.72rem;font-weight:700;background:' + bg + ';color:' + tc + ';padding:4px 14px;border-radius:20px;letter-spacing:.04em">' + lbl + '</span>'
        '</div>'
    )

def score_bar(label, score, max_score, icon=""):
    pct = (score / max_score * 100) if max_score else 0
    color = _bar_c(pct)
    return (
        '<div style="margin:14px 0">'
        '<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:7px">'
        '<span style="font-size:.855rem;font-weight:600;color:#C8C8F0">' + icon + '  ' + label + '</span>'
        '<div style="display:flex;align-items:baseline;gap:3px">'
        '<span style="font-size:1rem;font-weight:800;color:' + color + ';letter-spacing:-.02em">' + str(round(score,1)) + '</span>'
        '<span style="font-size:.72rem;color:#4A4A6A;font-weight:500">/ ' + str(int(max_score)) + '</span>'
        '</div></div>'
        '<div style="background:rgba(255,255,255,.06);border-radius:999px;height:9px;overflow:hidden">'
        '<div style="background:' + color + ';width:' + str(round(pct,1)) + '%;height:100%;border-radius:999px;'
        'box-shadow:0 0 10px ' + color + '66;transition:width .6s ease"></div>'
        '</div></div>'
    )

def issue_box(text):
    return ('<div style="display:flex;gap:10px;align-items:flex-start;'
            'background:rgba(234,179,8,.09);border:1px solid rgba(234,179,8,.28);'
            'border-left:3px solid #EAB308;border-radius:0 10px 10px 0;'
            'padding:11px 14px;margin:6px 0;font-size:.83rem;color:#FDE047;line-height:1.55">'
            '<span style="flex-shrink:0;font-size:.9rem">\u26a0\ufe0f</span>'
            '<span>' + text + '</span></div>')

def ok_box(label):
    return ('<div style="display:flex;gap:9px;align-items:center;'
            'background:rgba(34,197,94,.09);border:1px solid rgba(34,197,94,.28);'
            'border-left:3px solid #22C55E;border-radius:0 10px 10px 0;'
            'padding:11px 14px;margin:6px 0;font-size:.83rem;color:#86EFAC;font-weight:500">'
            '\u2713  ' + label + ' \u2014 no issues found.</div>')

def pill(text, t="found"):
    styles = {
        "found":       "background:rgba(34,197,94,.13);color:#86EFAC;border:1px solid rgba(34,197,94,.32)",
        "missing":     "background:rgba(239,68,68,.13);color:#FCA5A5;border:1px solid rgba(239,68,68,.32)",
        "suggest":     "background:rgba(124,58,237,.18);color:#C4B5FD;border:1px solid rgba(124,58,237,.38)",
        "title_miss":  "background:rgba(234,179,8,.12);color:#FDE047;border:1px solid rgba(234,179,8,.32)",
        "bullet_miss": "background:rgba(249,115,22,.12);color:#FED7AA;border:1px solid rgba(249,115,22,.32)",
    }
    icons = {"found":"\u2713","missing":"\u2717","suggest":"\u25B8","title_miss":"T","bullet_miss":"B"}
    s = styles.get(t, styles["found"]); i = icons.get(t,"")
    return ('<span style="' + s + ';display:inline-flex;align-items:center;gap:4px;border-radius:20px;'
            'padding:4px 11px;margin:3px;font-size:.775rem;font-weight:600;line-height:1.4">' + i + '  ' + text + '</span>')

def section_label(text):
    # FIXED: was #4A4A7A (2.24:1) — now #7B7BAA (4.6:1) ✅
    return ('<p style="font-size:.7rem;font-weight:700;text-transform:uppercase;'
            'letter-spacing:.09em;color:#7B7BAA;margin:0 0 10px">' + text + '</p>')

def copy_btn(text, uid):
    esc = text.replace("\\","\\\\").replace("'","\\'").replace("\n"," ")
    return ('<button onclick="navigator.clipboard.writeText(\'' + esc + '\');'
            'this.innerText=\'\u2713 Copied!\';setTimeout(()=>this.innerText=\'Copy\',2000)" '
            'style="background:rgba(124,58,237,.18);border:1px solid rgba(124,58,237,.4);'
            'border-radius:8px;padding:6px 16px;font-size:.775rem;font-weight:600;'
            'color:#C4B5FD;cursor:pointer;margin-top:10px;transition:all .15s;'
            'font-family:Inter,sans-serif">Copy</button>')

def confidence_badge(conf, strategy):
    if conf >= 85: c, bg = "#86EFAC","rgba(34,197,94,.13)"
    elif conf >= 75: c, bg = "#FDE047","rgba(234,179,8,.13)"
    else: c, bg = "#FED7AA","rgba(249,115,22,.13)"
    return ('<span style="background:' + bg + ';color:' + c + ';border-radius:8px;'
            'padding:3px 9px;font-size:.72rem;font-weight:700;margin-left:8px;'
            'border:1px solid ' + c + '44">'
            + str(conf) + '% \u00b7 ' + strategy + '</span>')

def page_header(title, sub=""):
    sub_html = ('<p style="color:#7B7BAA;font-size:.9rem;margin:5px 0 0;font-weight:400">' + sub + '</p>') if sub else ""
    show('<div style="margin-bottom:30px;padding-bottom:24px;border-bottom:1px solid rgba(124,58,237,.15)">'
         '<div style="display:flex;align-items:center;gap:14px;margin-bottom:4px">'
         '<div style="width:4px;height:32px;background:linear-gradient(180deg,#7C3AED,#C4B5FD);'
         'border-radius:4px;box-shadow:0 0 12px rgba(124,58,237,.5)"></div>'
         '<h1 style="font-size:1.7rem;font-weight:800;color:#F0F0FF;margin:0;letter-spacing:-.035em">' + title + '</h1>'
         '</div>' + sub_html + '</div>')

def empty_state(icon, title, subtitle):
    show('<div style="background:#10101E;border-radius:20px;padding:56px 24px;'
         'border:1px solid rgba(124,58,237,.15);text-align:center;margin:20px 0">'
         '<div style="font-size:3rem;margin-bottom:18px">' + icon + '</div>'
         '<p style="font-size:1.05rem;font-weight:700;color:#C8C8F0;margin:0 0 10px">' + title + '</p>'
         '<p style="font-size:.875rem;color:#7B7BAA;margin:0">' + subtitle + '</p></div>')

# ── SIDEBAR ─────────────────────────────────────────────────────────────────────
with st.sidebar:
    show('<div style="padding:10px 6px 22px">'
         '<div style="display:flex;align-items:center;gap:12px;margin-bottom:4px">'
         '<div style="width:36px;height:36px;background:linear-gradient(135deg,#6D28D9,#8B5CF6);'
         'border-radius:10px;display:flex;align-items:center;justify-content:center;font-size:1rem;'
         'box-shadow:0 4px 16px rgba(109,40,217,.45)">\U0001f6d2</div>'
         '<div>'
         '<div style="font-size:.95rem;font-weight:800;color:#EDE9FE;letter-spacing:-.02em">Amazon</div>'
         '<div style="font-size:.95rem;font-weight:800;color:#8B5CF6;letter-spacing:-.02em">Analyser</div>'
         '</div></div>'
         '<p style="font-size:.68rem;color:#5A5A80;margin:8px 0 0;letter-spacing:.04em">CONTENT \u00b7 SEO \u00b7 UK</p>'
         '</div>')
    st.markdown("---")
    page = st.radio("Navigation", [
        "\U0001f3e0  Dashboard",
        "\U0001f50d  New Analysis",
        "\U0001f4dc  History",
        "\U0001f3f7  Product Lines",
        "\u2699\ufe0f  Scoring Rules",
    ], label_visibility="collapsed")
    st.markdown("---")
    show('<div style="padding:0 6px">'
         + section_label("Score guide") +
         "<div style='display:flex;flex-direction:column;gap:7px'>" +
         "".join(
             '<div style="display:flex;align-items:center;gap:10px">'
             '<div style="width:9px;height:9px;border-radius:50%;background:' + c + ';'
             'box-shadow:0 0 7px ' + c + ';flex-shrink:0"></div>'
             '<span style="font-size:.82rem;color:#A0A0D0;font-weight:500">' + lbl + '</span>'
             '<span style="font-size:.78rem;color:#5A5A80;margin-left:auto">' + rng + '</span>'
             '</div>'
             for c, lbl, rng in [
                 ("#22C55E","Excellent","\u226585"),
                 ("#EAB308","Good","65\u201384"),
                 ("#F97316","Needs Work","40\u201364"),
                 ("#EF4444","Poor","<40"),
             ]
         ) + '</div></div>')


# ── RESULT CARD ────────────────────────────────────────────────────────────────
def render_result(asin, title, bullets_json, has_aplus, image,
                  total, t_s, b_s, a_s, k_s,
                  t_issues, b_issues, a_issues, k_issues,
                  found_kws, miss_kws, sugg_kws, rule_w, pl_keywords):
    s = total or 0
    bullets = json.loads(bullets_json or "[]")
    t_miss, b_miss = keyword_gaps(title, bullets, pl_keywords)
    rewrites = bullet_rewrite_suggestions(bullets, b_miss) if bullets else []
    c_, bg_, tc_, lbl_ = _score_meta(s)
    img_h = ('<img src="' + (image or "") + '" style="width:76px;height:76px;object-fit:contain;'
             'border-radius:12px;border:1px solid rgba(124,58,237,.2);margin-bottom:14px;'
             'background:#0F0F20">') if image else ""
    bars = (score_bar("Title",        t_s, rule_w.get("title_weight",25),   "\U0001f4dd") +
            score_bar("Bullet Points",b_s, rule_w.get("bullets_weight",25), "\U0001f539") +
            score_bar("A+ Content",   a_s, rule_w.get("aplus_weight",25),   "\u2728")    +
            score_bar("Keywords SEO", k_s, rule_w.get("keywords_weight",25),"\U0001f511"))
    all_issues = t_issues + b_issues + a_issues + k_issues
    issues_html = "".join(issue_box(i) for i in all_issues) if all_issues else ok_box("All sections")
    header = (
        '<div style="display:flex;justify-content:space-between;align-items:flex-start;'
        'gap:24px;margin-bottom:26px;flex-wrap:wrap">'
        '<div style="flex:1;min-width:220px">' + img_h +
        '<div style="display:flex;align-items:center;gap:10px;margin-bottom:12px;flex-wrap:wrap">'
        '<code style="background:rgba(124,58,237,.22);color:#C4B5FD;padding:5px 13px;'
        'border-radius:8px;font-size:.85rem;font-weight:700;letter-spacing:.05em">' + asin + '</code>'
        '<span style="background:' + bg_ + ';color:' + tc_ + ';padding:4px 13px;border-radius:20px;'
        'font-size:.75rem;font-weight:700;letter-spacing:.03em">' + lbl_ + '</span>'
        '<a href="https://www.amazon.co.uk/dp/' + asin + '" target="_blank" '
        'style="font-size:.78rem;color:#7B7BAA;text-decoration:none;'
        'padding:4px 10px;border:1px solid rgba(255,255,255,.1);border-radius:8px;'
        'transition:all .15s">\U0001f517 View on Amazon</a>'
        '</div>'
        '<p style="color:#A0A0D0;font-size:.88rem;margin:0;line-height:1.65;'
        'border-left:2px solid rgba(124,58,237,.3);padding-left:12px">'
        + (title or "\u2014")[:140] + '</p>'
        '</div>' + score_gauge(s) + '</div>'
    )
    show('<div style="background:#10101E;border-radius:22px;padding:28px 32px;'
         'border:1px solid rgba(124,58,237,.18);margin:20px 0;'
         'box-shadow:0 8px 40px rgba(0,0,0,.4),0 0 0 1px rgba(124,58,237,.06)">'
         + header +
         '<div style="border-top:1px solid rgba(255,255,255,.05);padding-top:20px;margin-top:4px">'
         + bars + '</div>'
         '<div style="margin-top:16px">' + issues_html + '</div>')
    if pl_keywords:
        show('<div style="margin-top:20px">' + section_label("Keyword gaps by section"))
        ghtml = ('<div style="display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-bottom:16px">'
                 '<div style="background:rgba(234,179,8,.07);border:1px solid rgba(234,179,8,.22);'
                 'border-radius:14px;padding:16px 18px">'
                 '<p style="font-size:.7rem;font-weight:700;text-transform:uppercase;letter-spacing:.08em;'
                 'color:#EAB308;margin:0 0 10px">\U0001f4dd Missing from Title</p>')
        ghtml += ("".join(pill(k,"title_miss") for k in t_miss) if t_miss
                  else '<span style="font-size:.82rem;color:#22C55E;font-weight:600">\u2713 All keywords present</span>')
        ghtml += ('</div>'
                  '<div style="background:rgba(249,115,22,.07);border:1px solid rgba(249,115,22,.22);'
                  'border-radius:14px;padding:16px 18px">'
                  '<p style="font-size:.7rem;font-weight:700;text-transform:uppercase;letter-spacing:.08em;'
                  'color:#F97316;margin:0 0 10px">\U0001f539 Missing from Bullet Points</p>')
        ghtml += ("".join(pill(k,"bullet_miss") for k in b_miss) if b_miss
                  else '<span style="font-size:.82rem;color:#22C55E;font-weight:600">\u2713 All keywords present</span>')
        ghtml += '</div></div>'
        show(ghtml + '</div>')
    if found_kws or miss_kws:
        show('<div style="margin-top:16px">' + section_label("Full keyword coverage"))
        show("".join(pill(k,"found") for k in found_kws) + "".join(pill(k,"missing") for k in miss_kws) +
             '<p style="font-size:.72rem;color:#5A5A80;margin:8px 0 0">'
             '\u2713 found in listing \u00b7 \u2717 not found anywhere</p></div>')
    if rewrites:
        show('<div style="margin-top:22px;border-top:1px solid rgba(255,255,255,.05);padding-top:20px">'
             + section_label("\U0001f4a1 Smart rewrite suggestions — add missing keywords to bullet points"))
        for rw in rewrites:
            uid = re.sub(r"\W","_", asin + "_" + rw["keyword"])
            show('<div style="background:#0F0F20;border:1px solid rgba(124,58,237,.18);'
                 'border-radius:14px;padding:18px 22px;margin-bottom:14px">'
                 '<div style="display:flex;align-items:center;flex-wrap:wrap;gap:6px;margin-bottom:14px">'
                 '<span style="font-size:.7rem;font-weight:700;text-transform:uppercase;letter-spacing:.07em;color:#7B7BAA">Keyword</span>'
                 '<span style="background:rgba(124,58,237,.2);color:#C4B5FD;border-radius:8px;'
                 'padding:3px 10px;font-size:.82rem;font-weight:700">' + rw["keyword"] + '</span>'
                 '<span style="color:#5A5A80;font-size:.8rem">\u2192 Bullet ' + str(rw["bullet_idx"]+1) + '</span>'
                 '<span style="background:rgba(255,255,255,.05);color:#7B7BAA;border-radius:6px;'
                 'padding:2px 8px;font-size:.72rem;border:1px solid rgba(255,255,255,.08)">type: ' + rw["bullet_type"] + '</span>'
                 + confidence_badge(rw["confidence"], rw["strategy"]) + '</div>'
                 '<div style="display:grid;grid-template-columns:1fr 1fr;gap:14px">'
                 '<div>'
                 '<p style="font-size:.68rem;font-weight:700;text-transform:uppercase;letter-spacing:.08em;'
                 'color:#5A5A80;margin:0 0 8px">Original Bullet</p>'
                 # FIXED: was #5050A0 (2.76:1) — now #8888C0 (5.55:1) ✅
                 '<p style="font-size:.84rem;color:#8888C0;line-height:1.6;margin:0;font-style:italic;'
                 'padding:12px 14px;background:rgba(255,255,255,.03);border-radius:10px;'
                 'border:1px solid rgba(255,255,255,.07)">' + rw["original"] + '</p>'
                 '</div>'
                 '<div>'
                 '<p style="font-size:.68rem;font-weight:700;text-transform:uppercase;letter-spacing:.08em;'
                 'color:#22C55E;margin:0 0 8px">Suggested Rewrite</p>'
                 '<p style="font-size:.84rem;color:#E8E8FF;line-height:1.6;margin:0;'
                 'padding:12px 14px;background:rgba(34,197,94,.07);'
                 'border:1px solid rgba(34,197,94,.22);border-radius:10px">' + rw["suggested"] + '</p>'
                 + copy_btn(rw["suggested"], uid) +
                 '</div></div></div>')
        show('</div>')
    if sugg_kws:
        show('<div style="margin-top:16px">' + section_label("Suggested keywords from page content"))
        show("".join(pill(k,"suggest") for k in sugg_kws) + '</div>')
    show('</div>')

# ── DASHBOARD ──────────────────────────────────────────────────────────────────
if page == "\U0001f3e0  Dashboard":
    page_header("Dashboard", "Your Amazon listing health at a glance")
    conn = db()
    rows   = conn.execute("SELECT * FROM results ORDER BY created_at DESC LIMIT 200").fetchall()
    plines = conn.execute("SELECT COUNT(*) FROM product_lines").fetchone()[0]
    conn.close()
    total_a = len(rows)
    avg_s   = round(sum(r["total_score"] or 0 for r in rows) / max(total_a,1), 1)
    poor    = sum(1 for r in rows if (r["total_score"] or 0) < 65)
    excellent = sum(1 for r in rows if (r["total_score"] or 0) >= 85)
    if total_a == 0:
        empty_state("\U0001f4e6","No analyses yet","Create a Product Line and run your first analysis to see your dashboard.")
    else:
        c1,c2,c3,c4 = st.columns(4)
        c1.metric("Total Analyses", total_a)
        c2.metric("Average Score", str(avg_s) + " / 100")
        c3.metric("Need Attention", poor, delta=(str(-poor) + " listings below 65") if poor else "All good \u2713")
        c4.metric("Product Lines", plines)
        show('<div style="height:24px"></div>')
        show('<h2 style="font-size:1.05rem;font-weight:700;color:#C8C8F0;margin:0 0 14px;letter-spacing:-.01em">Recent Analyses</h2>')
        tbl = ('<div style="background:#10101E;border-radius:20px;border:1px solid rgba(124,58,237,.15);overflow:hidden">'
               '<table style="width:100%;border-collapse:collapse;font-size:.84rem">'
               '<thead><tr style="background:rgba(124,58,237,.1)">')
        for h in ["ASIN","Product","Score","Title","Bullets","A+","Keywords","Date"]:
            tbl += ('<th style="padding:12px 16px;text-align:left;font-size:.68rem;font-weight:700;'
                    'text-transform:uppercase;letter-spacing:.09em;color:#7B7BAA;white-space:nowrap;'
                    'border-bottom:1px solid rgba(124,58,237,.15)">' + h + '</th>')
        tbl += '</tr></thead><tbody>'
        for r in rows[:15]:
            s = r["total_score"]; c_,bg_,tc_,lbl_ = _score_meta(s)
            sc = ('<span style="background:' + bg_ + ';color:' + tc_ + ';padding:3px 11px;border-radius:20px;'
                  'font-size:.75rem;font-weight:700">' + str(int(s)) + ' \u2014 ' + lbl_ + '</span>') if s is not None else '<span style="color:#5A5A80">Error</span>'
            ap = ('<span style="color:#22C55E;font-weight:700;font-size:.82rem">\u2713 Yes</span>'
                  if r["has_aplus"] else '<span style="color:#EF4444;font-weight:700;font-size:.82rem">\u2717 No</span>')
            fv = lambda v: str(int(v)) if v else "\u2014"
            tbl += ('<tr style="border-bottom:1px solid rgba(255,255,255,.04);transition:background .1s">'
                    '<td style="padding:12px 16px"><code>' + r["asin"] + '</code></td>'
                    # FIXED: was #5050A0 (2.63:1) — now #8888C0 ✅
                    '<td style="padding:12px 16px;color:#8888C0;max-width:170px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">'
                    + (r["product_name"] or "\u2014")[:45] + '</td>'
                    '<td style="padding:12px 16px">' + sc + '</td>'
                    '<td style="padding:12px 16px;color:#9090C0">' + fv(r["title_score"]) + '</td>'
                    '<td style="padding:12px 16px;color:#9090C0">' + fv(r["bullets_score"]) + '</td>'
                    '<td style="padding:12px 16px">' + ap + '</td>'
                    '<td style="padding:12px 16px;color:#9090C0">' + fv(r["keywords_score"]) + '</td>'
                    '<td style="padding:12px 16px;color:#5A5A80;font-size:.78rem">' + str(r["created_at"] or "")[:10] + '</td>'
                    '</tr>')
        tbl += '</tbody></table></div>'
        show(tbl)

# ── NEW ANALYSIS ───────────────────────────────────────────────────────────────
elif page == "\U0001f50d  New Analysis":
    page_header("New Analysis", "Scrape and score Amazon.co.uk listings in seconds")
    conn = db()
    lines = conn.execute("SELECT * FROM product_lines ORDER BY name").fetchall()
    rules = conn.execute("SELECT * FROM scoring_rules ORDER BY is_default DESC, name").fetchall()
    conn.close()
    if not lines:
        empty_state("\U0001f3f7\ufe0f","No product lines yet",
                    "Go to Product Lines \u2192 create one with your target keywords \u2192 come back here.")
        st.stop()
    show('<div style="background:#10101E;border-radius:18px;padding:24px 28px;'
         'border:1px solid rgba(124,58,237,.18);margin-bottom:20px">')
    col_a, col_b, col_c = st.columns([2,1,1], gap="large")
    with col_a:
        show(section_label("ASINs \u2014 one per line"))
        asins_raw = st.text_area("asins","", placeholder="B08N5WRWNW\nB09XYZ1234\nB07ABCDEF1",
                                  height=130, label_visibility="collapsed")
    with col_b:
        show(section_label("Product Line"))
        line_names = [l["name"] + "  (" + str(len(json.loads(l["keywords"]))) + " kws)" for l in lines]
        sel_line = st.selectbox("line","" , options=line_names, label_visibility="collapsed")
    with col_c:
        show(section_label("Scoring Rule"))
        rule_names = [r["name"] + (" \u2605" if r["is_default"] else "") for r in rules]
        sel_rule = st.selectbox("rule","", options=rule_names, label_visibility="collapsed")
    show('</div>')
    if st.button("\u25b6\ufe0f  Run Analysis", type="primary", use_container_width=True):
        asins = [a.strip().upper() for a in asins_raw.strip().splitlines() if a.strip()]
        if not asins: st.error("Please enter at least one ASIN."); st.stop()
        line_obj  = lines[line_names.index(sel_line)]
        rule_obj  = rules[rule_names.index(sel_rule)]
        keywords  = json.loads(line_obj["keywords"])
        rule_dict = dict(rule_obj)
        bar = st.progress(0, text="Initialising\u2026")
        for idx, asin in enumerate(asins):
            bar.progress(idx / max(len(asins),1),
                         text="\U0001f50d Scraping " + asin + " (" + str(idx+1) + "/" + str(len(asins)) + ")\u2026")
            conn = db()
            try:
                scraped  = scrape(asin)
                result   = analyse(scraped, keywords, rule_dict)
                pname    = scraped.get("brand") or (scraped.get("title") or "")[:60] or asin
                conn.execute(
                    "INSERT INTO results (asin,product_line_id,scoring_rule_id,product_name,product_image,"
                    "title,bullets,has_aplus,description,total_score,title_score,bullets_score,aplus_score,"
                    "keywords_score,title_issues,bullets_issues,aplus_issues,keywords_issues,"
                    "found_keywords,missing_keywords,suggested_keywords) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                    (asin,line_obj["id"],rule_obj["id"],pname,scraped.get("image_url"),
                     scraped.get("title"),json.dumps(scraped.get("bullets",[])),
                     1 if scraped.get("has_aplus") else 0,scraped.get("description"),
                     result["total_score"],result["title_score"],result["bullets_score"],
                     result["aplus_score"],result["keywords_score"],
                     json.dumps(result["title_issues"]),json.dumps(result["bullets_issues"]),
                     json.dumps(result["aplus_issues"]),json.dumps(result["keywords_issues"]),
                     json.dumps(result["found_keywords"]),json.dumps(result["missing_keywords"]),
                     json.dumps(result["suggested_keywords"])))
                conn.commit()
                render_result(asin, scraped.get("title"), json.dumps(scraped.get("bullets",[])),
                              scraped.get("has_aplus"), scraped.get("image_url"),
                              result["total_score"], result["title_score"], result["bullets_score"],
                              result["aplus_score"], result["keywords_score"],
                              result["title_issues"], result["bullets_issues"],
                              result["aplus_issues"], result["keywords_issues"],
                              result["found_keywords"], result["missing_keywords"], result["suggested_keywords"],
                              rule_dict, keywords)
            except Exception as e:
                conn.execute("INSERT INTO results (asin,product_line_id,scoring_rule_id,scrape_error,total_score) VALUES (?,?,?,?,0)",
                             (asin, line_obj["id"], rule_obj["id"], str(e)))
                conn.commit()
                show('<div style="background:rgba(239,68,68,.1);border:1px solid rgba(239,68,68,.3);'
                     'border-left:3px solid #EF4444;border-radius:0 10px 10px 0;padding:12px 16px;margin:8px 0">'
                     '<span style="color:#FCA5A5;font-weight:600">' + asin + '</span>'
                     '<span style="color:#7B7BAA;font-size:.85rem"> \u2014 ' + str(e) + '</span></div>')
            finally:
                conn.close()
        bar.progress(1.0, text="\u2713 Analysis complete!")

# ── HISTORY ────────────────────────────────────────────────────────────────────
elif page == "\U0001f4dc  History":
    page_header("History", "All past analyses with full details")
    conn = db()
    rows = conn.execute(
        "SELECT r.*, pl.keywords as pl_keywords FROM results r "
        "LEFT JOIN product_lines pl ON r.product_line_id=pl.id ORDER BY r.created_at DESC").fetchall()
    rules_map = {r["id"]: dict(r) for r in conn.execute("SELECT * FROM scoring_rules").fetchall()}
    conn.close()
    if not rows:
        empty_state("\U0001f4dc","No history yet","Run your first analysis and all results will appear here.")
    else:
        hc1, hc2 = st.columns([3,1])
        hc1.caption(str(len(rows)) + " result" + ("s" if len(rows)!=1 else "") + " stored")
        xl = build_excel(rows)
        hc2.download_button("\u2193  Export to Excel", data=xl,
                            file_name="amazon_analysis.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True)
        show('<div style="height:12px"></div>')
        for r in rows:
            s = r["total_score"]; c_,bg_,tc_,lbl_ = _score_meta(s)
            score_str = str(int(s)) + "/100  " + lbl_ if s is not None else "Error"
            exp_label = (r["asin"] + "  \u2014  " + score_str + "  \u00b7  "
                         + (r["product_name"] or "\u2014")[:38]
                         + "  \u00b7  " + str(r["created_at"] or "")[:10])
            with st.expander(exp_label):
                if r["scrape_error"]:
                    show('<div style="background:rgba(239,68,68,.1);border:1px solid rgba(239,68,68,.3);'
                         'border-radius:10px;padding:12px 16px;margin-bottom:12px">'
                         '<p style="color:#FCA5A5;font-weight:600;margin:0 0 4px">Scraping failed</p>'
                         '<p style="color:#7B7BAA;font-size:.85rem;margin:0">' + r["scrape_error"] + '</p></div>')
                    c2 = db()
                    if st.button("\U0001f5d1  Delete", key="del_" + str(r["id"])):
                        c2.execute("DELETE FROM results WHERE id=?", (r["id"],)); c2.commit(); st.rerun()
                    c2.close(); continue
                pl_kws  = json.loads(r["pl_keywords"] or "[]") if "pl_keywords" in r.keys() else []
                rule_w  = rules_map.get(r["scoring_rule_id"],
                                        {"title_weight":25,"bullets_weight":25,"aplus_weight":25,"keywords_weight":25})
                render_result(r["asin"], r["title"], r["bullets"], r["has_aplus"], r["product_image"],
                              r["total_score"], r["title_score"], r["bullets_score"],
                              r["aplus_score"], r["keywords_score"],
                              json.loads(r["title_issues"] or "[]"), json.loads(r["bullets_issues"] or "[]"),
                              json.loads(r["aplus_issues"] or "[]"), json.loads(r["keywords_issues"] or "[]"),
                              json.loads(r["found_keywords"] or "[]"), json.loads(r["missing_keywords"] or "[]"),
                              json.loads(r["suggested_keywords"] or "[]"), rule_w, pl_kws)
                c2 = db()
                if st.button("\U0001f5d1  Delete result", key="del2_" + str(r["id"])):
                    c2.execute("DELETE FROM results WHERE id=?", (r["id"],)); c2.commit(); st.rerun()
                c2.close()

# ── PRODUCT LINES ──────────────────────────────────────────────────────────────
elif page == "\U0001f3f7  Product Lines":
    page_header("Product Lines","Group products and assign target keyword lists")
    conn = db()
    lines = conn.execute("SELECT * FROM product_lines ORDER BY name").fetchall()
    with st.expander("\u2795  Create new product line", expanded=len(lines)==0):
        c1, c2 = st.columns([1,1], gap="large")
        with c1:
            new_name = st.text_input("Product line name", placeholder="e.g. Coffee Machines")
        new_kws = st.text_area("Keywords",
                               placeholder="One keyword per line or comma-separated:\ncoffee machine\nespresso maker\nautomatic coffee maker\nbarista",
                               height=190)
        kc = len([k.strip() for k in re.split(r"[,\n]+", new_kws) if k.strip()])
        ca, cb = st.columns([1,4])
        ca.caption(str(kc) + " keywords detected")
        if cb.button("Create product line", type="primary", key="cpl"):
            if not new_name.strip():
                st.error("Please enter a name.")
            else:
                kws = [k.strip() for k in re.split(r"[,\n]+", new_kws) if k.strip()]
                try:
                    conn.execute("INSERT INTO product_lines (name,keywords) VALUES (?,?)",
                                 (new_name.strip(), json.dumps(kws)))
                    conn.commit()
                    st.success("\u2713 Product line '" + new_name.strip() + "' created with " + str(len(kws)) + " keywords!")
                    st.rerun()
                except Exception as e: st.error(str(e))
    show('<div style="height:10px"></div>')
    if not lines:
        empty_state("\U0001f3f7\ufe0f","No product lines yet","Use the form above to create your first product line.")
    for line in lines:
        kws = json.loads(line["keywords"])
        with st.expander(line["name"] + "  \u00b7  " + str(len(kws)) + " keywords  \u00b7  created " + str(line["created_at"] or "")[:10]):
            show("".join(pill(k,"suggest") for k in kws[:30])
                 + (' <span style="color:#7B7BAA;font-size:.78rem">+' + str(len(kws)-30) + ' more</span>' if len(kws)>30 else ""))
            show('<div style="height:10px"></div>')
            c1, c2 = st.columns([1,1], gap="large")
            with c1: en = st.text_input("Name", value=line["name"], key="ln_"+str(line["id"]))
            ek = st.text_area("Keywords (one per line)", value="\n".join(kws), height=190, key="lk_"+str(line["id"]))
            ca, cb = st.columns([1,4])
            if ca.button("Save", key="ls_"+str(line["id"])):
                nk = [k.strip() for k in ek.strip().splitlines() if k.strip()]
                conn.execute("UPDATE product_lines SET name=?,keywords=? WHERE id=?",
                             (en.strip(), json.dumps(nk), line["id"]))
                conn.commit(); st.success("\u2713 Saved!"); st.rerun()
            if cb.button("Delete this product line", key="ld_"+str(line["id"])):
                conn.execute("DELETE FROM product_lines WHERE id=?", (line["id"],))
                conn.commit(); st.rerun()
    conn.close()

# ── SCORING RULES ──────────────────────────────────────────────────────────────
elif page == "\u2699\ufe0f  Scoring Rules":
    page_header("Scoring Rules","Define how listings are evaluated and weighted")
    conn = db()
    rules = conn.execute("SELECT * FROM scoring_rules ORDER BY is_default DESC, name").fetchall()
    with st.expander("\u2795  Create new scoring rule"):
        rname = st.text_input("Rule name", placeholder="e.g. Strict SEO, Premium Listings")
        show(section_label("Category weights \u2014 must sum to 100"))
        rc1,rc2,rc3,rc4 = st.columns(4)
        tw  = rc1.number_input("Title %",     0, 100, 25, key="ntw")
        bw  = rc2.number_input("Bullets %",   0, 100, 25, key="nbw")
        aw  = rc3.number_input("A+ Content %",0, 100, 25, key="naw")
        kw_ = rc4.number_input("Keywords %",  0, 100, 25, key="nkw")
        tw_ = tw + bw + aw + kw_; ok = tw_ == 100
        show('<div style="background:' + ("rgba(34,197,94,.08)" if ok else "rgba(239,68,68,.08)") +
             ';border:1px solid ' + ("rgba(34,197,94,.25)" if ok else "rgba(239,68,68,.25)") +
             ';border-radius:10px;padding:10px 14px;margin:8px 0 16px;font-size:.85rem;font-weight:600;color:'
             + ("#86EFAC" if ok else "#FCA5A5") + '">'
             + ("\u2713 Weights sum to 100 \u2014 good to go" if ok else "\u26a0\ufe0f Total is " + str(tw_) + "/100 \u2014 must equal exactly 100")
             + '</div>')
        cx, cy = st.columns(2, gap="large")
        with cx:
            show(section_label("Title rules"))
            tca, tcb = st.columns(2)
            tmin = tca.number_input("Min chars",  1, 500, 80,  key="ntmin")
            tmax = tcb.number_input("Max chars",  1, 500, 200, key="ntmax")
            tkif = st.checkbox("Primary keyword must appear in first 80 chars", True, key="ntkif")
        with cy:
            show(section_label("Bullet point rules"))
            bca, bcb, bcc = st.columns(3)
            bmc = bca.number_input("Min count", 1, 20,   5,   key="nbmc")
            bml = bcb.number_input("Min chars", 1, 500,  100, key="nbml")
            bmx = bcc.number_input("Max chars", 1, 1000, 255, key="nbmx")
            show(section_label("Keyword rules"))
            kmc = st.slider("Minimum keyword coverage %", 0, 100, 70, key="nkmc")
            kit = st.checkbox("Top 3 keywords must appear in title", True, key="nkit")
        if st.button("Create rule", type="primary", disabled=(not ok), key="cr"):
            if not rname.strip(): st.error("Enter a rule name.")
            else:
                try:
                    conn.execute(
                        "INSERT INTO scoring_rules (name,title_weight,bullets_weight,aplus_weight,keywords_weight,"
                        "title_min_length,title_max_length,title_keyword_in_first,bullets_min_count,"
                        "bullets_min_length,bullets_max_length,keywords_min_coverage,keywords_in_title) "
                        "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
                        (rname.strip(),tw,bw,aw,kw_,tmin,tmax,int(tkif),bmc,bml,bmx,kmc,int(kit)))
                    conn.commit(); st.success("\u2713 Rule created!"); st.rerun()
                except Exception as e: st.error(str(e))
    show('<div style="height:10px"></div>')
    for rule in rules:
        tag = " \u2605 Default" if rule["is_default"] else ""
        with st.expander(rule["name"] + tag):
            show('<div style="display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-bottom:10px">'
                 '<div style="background:#0F0F20;border:1px solid rgba(124,58,237,.15);border-radius:12px;padding:16px 18px">'
                 + section_label("Weights") +
                 '<div style="display:flex;flex-direction:column;gap:6px">'
                 + "".join('<div style="display:flex;justify-content:space-between;align-items:center">'
                           '<span style="font-size:.84rem;color:#A0A0D0">' + n + '</span>'
                           '<span style="font-size:.9rem;font-weight:700;color:#C4B5FD">' + str(v) + '%</span>'
                           '</div>'
                           for n, v in [("Title", rule["title_weight"]),("Bullets",rule["bullets_weight"]),
                                        ("A+ Content",rule["aplus_weight"]),("Keywords",rule["keywords_weight"])])
                 + '</div></div>'
                 '<div style="background:#0F0F20;border:1px solid rgba(124,58,237,.15);border-radius:12px;padding:16px 18px">'
                 + section_label("Thresholds") +
                 '<div style="display:flex;flex-direction:column;gap:6px">'
                 + "".join('<div style="display:flex;justify-content:space-between;align-items:center">'
                           '<span style="font-size:.84rem;color:#A0A0D0">' + n + '</span>'
                           '<span style="font-size:.84rem;font-weight:600;color:#C4B5FD">' + v + '</span>'
                           '</div>'
                           for n, v in [
                               ("Title length", str(rule["title_min_length"])+"\u2013"+str(rule["title_max_length"])+" chars"),
                               ("Min bullets", str(rule["bullets_min_count"])),
                               ("Bullet length", str(rule["bullets_min_length"])+"\u2013"+str(rule["bullets_max_length"])+" chars"),
                               ("KW coverage", "\u2265" + str(round(rule["keywords_min_coverage"])) + "%"),
                           ])
                 + '</div></div></div>')
            if not rule["is_default"]:
                if st.button("\U0001f5d1  Delete rule", key="rd_"+str(rule["id"])):
                    conn.execute("DELETE FROM scoring_rules WHERE id=?", (rule["id"],))
                    conn.commit(); st.rerun()
    conn.close()
